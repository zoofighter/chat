#!/usr/bin/env python3
"""
Phase 3, 4 완전 구현 스크립트
PDF/Excel 내보내기 + ALMAgent 확장 + 도구/프롬프트 업데이트
"""

import json
import sys

NOTEBOOK_PATH = "chatbot.ipynb"

print("="*70)
print("Phase 3, 4 완전 구현 시작")
print("Report-Agent 최종 완성!")
print("="*70)

# 노트북 로드
with open(NOTEBOOK_PATH, 'r') as f:
    nb = json.load(f)

print(f"\n✓ 현재 노트북: {len(nb['cells'])}개 셀")

def find_cell_by_content(nb, search_text):
    for idx, cell in enumerate(nb['cells']):
        if cell['cell_type'] == 'code':
            source = ''.join(cell.get('source', []))
            if search_text in source:
                return idx
    return None

def insert_cell_after(nb, after_idx, cell_type, source):
    new_cell = {
        "cell_type": cell_type,
        "metadata": {},
        "source": source if isinstance(source, list) else [source]
    }
    if cell_type == "code":
        new_cell["execution_count"] = None
        new_cell["outputs"] = []
    nb['cells'].insert(after_idx + 1, new_cell)
    return after_idx + 1

def replace_cell_source(nb, cell_idx, new_source):
    nb['cells'][cell_idx]['source'] = new_source if isinstance(new_source, list) else [new_source]

print("\n" + "="*70)
print("Phase 3: PDF/Excel 내보내기 함수 추가")
print("="*70)

# Phase 3: export_to_markdown, export_to_pdf, export_to_excel, export_report 추가
print("\nStep 1: export_to_markdown() 함수 추가...")

export_markdown_idx = find_cell_by_content(nb, 'def export_to_markdown')
if not export_markdown_idx:
    # Cell 10 찾기 (generate_comprehensive_report 이후)
    report_func_idx = find_cell_by_content(nb, 'def generate_comprehensive_report')

    export_markdown_code = '''def export_to_markdown(report_data: Dict[str, Any], output_path: str) -> str:
    """
    리포트를 Markdown 형식으로 내보내기 (Phase 1)

    Args:
        report_data: generate_comprehensive_report()로 생성된 리포트 데이터
        output_path: 저장할 파일 경로

    Returns:
        저장된 파일 경로
    """
    import os

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)

    lines = []

    # 제목
    lines.append(f"# {report_data['title']}\\n")
    lines.append(f"생성일시: {report_data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')}\\n")

    if report_data['metadata'].get('scenario_no'):
        lines.append(f"시나리오: {report_data['metadata']['scenario_no']}\\n")

    lines.append("\\n---\\n\\n")

    # 각 섹션
    sections = report_data.get('sections', {})

    for section_name, section_data in sections.items():
        lines.append(f"## {section_data['title']}\\n\\n")
        lines.append(f"{section_data.get('summary', '')}\\n\\n")

        # 데이터 테이블
        if 'data' in section_data and section_data['data']:
            data = section_data['data'][:20]  # 최대 20행
            if data:
                # 헤더
                headers = list(data[0].keys())
                lines.append("| " + " | ".join(headers) + " |\\n")
                lines.append("| " + " | ".join(['---'] * len(headers)) + " |\\n")

                # 데이터 행
                for row in data:
                    values = [str(row.get(h, '')) for h in headers]
                    lines.append("| " + " | ".join(values) + " |\\n")

                lines.append("\\n")

        lines.append("\\n")

    # 파일 저장
    with open(output_path, 'w', encoding='utf-8') as f:
        f.writelines(lines)

    return output_path

print("✓ Markdown 내보내기 함수 정의 완료! (Phase 1)")
'''

    insert_cell_after(nb, report_func_idx, 'code', export_markdown_code)
    print("  ✓ export_to_markdown() 함수 추가 완료")

print("\nStep 2: export_to_pdf() 함수 추가...")

export_pdf_code = '''def export_to_pdf(report_data: Dict[str, Any], output_path: str) -> str:
    """
    리포트를 PDF 형식으로 내보내기 (Phase 3)

    Args:
        report_data: generate_comprehensive_report()로 생성된 리포트 데이터
        output_path: 저장할 파일 경로

    Returns:
        저장된 파일 경로
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm
    import os

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)

    doc = SimpleDocTemplate(output_path, pagesize=A4)
    story = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30
    )
    heading_style = styles['Heading2']

    # 제목
    story.append(Paragraph(report_data['title'], title_style))
    story.append(Spacer(1, 0.5*cm))

    # 메타데이터
    meta_text = f"생성일시: {report_data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')}"
    if report_data['metadata'].get('scenario_no'):
        meta_text += f" | 시나리오: {report_data['metadata']['scenario_no']}"
    story.append(Paragraph(meta_text, styles['Normal']))
    story.append(Spacer(1, 1*cm))

    # 각 섹션
    sections = report_data.get('sections', {})

    for section_name, section_data in sections.items():
        story.append(Paragraph(section_data['title'], heading_style))
        story.append(Spacer(1, 0.3*cm))

        story.append(Paragraph(section_data.get('summary', ''), styles['Normal']))
        story.append(Spacer(1, 0.5*cm))

        if 'data' in section_data and section_data['data']:
            data = section_data['data'][:10]
            if data:
                headers = list(data[0].keys())
                table_data = [headers]

                for row in data:
                    table_data.append([str(row.get(h, '')) for h in headers])

                t = Table(table_data)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))

                story.append(t)
                story.append(Spacer(1, 0.5*cm))

        story.append(Spacer(1, 0.5*cm))

    doc.build(story)

    return output_path

print("✓ PDF 내보내기 함수 정의 완료! (Phase 3)")
'''

markdown_func_idx = find_cell_by_content(nb, 'def export_to_markdown')
if markdown_func_idx:
    insert_cell_after(nb, markdown_func_idx, 'code', export_pdf_code)
    print("  ✓ export_to_pdf() 함수 추가 완료")

print("\nStep 3: export_to_excel() 함수 추가...")

export_excel_code = '''def export_to_excel(report_data: Dict[str, Any], output_path: str) -> str:
    """
    리포트를 Excel 형식으로 내보내기 (Phase 3)

    Args:
        report_data: generate_comprehensive_report()로 생성된 리포트 데이터
        output_path: 저장할 파일 경로

    Returns:
        저장된 파일 경로
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    import os

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)

    wb = Workbook()

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # 요약 시트
    summary_ws = wb.create_sheet(title="요약")
    summary_ws.append([report_data['title']])
    summary_ws.append([])
    summary_ws.append(['생성일시', report_data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')])

    if report_data['metadata'].get('scenario_no'):
        summary_ws.append(['시나리오', report_data['metadata']['scenario_no']])

    summary_ws.append([])
    summary_ws.append(['섹션', '요약'])

    # 스타일
    title_font = Font(size=16, bold=True)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    summary_ws['A1'].font = title_font

    # 각 섹션
    sections = report_data.get('sections', {})

    for section_name, section_data in sections.items():
        summary_ws.append([section_data['title'], section_data.get('summary', '')])

        ws = wb.create_sheet(title=section_data['title'][:30])

        ws.append([section_data['title']])
        ws.append([section_data.get('summary', '')])
        ws.append([])

        if 'data' in section_data and section_data['data']:
            data = section_data['data']
            if data:
                headers = list(data[0].keys())
                ws.append(headers)

                for col_idx, _ in enumerate(headers, 1):
                    cell = ws.cell(row=ws.max_row, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')

                for row_data in data:
                    ws.append([row_data.get(h, '') for h in headers])

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)

    return output_path

print("✓ Excel 내보내기 함수 정의 완료! (Phase 3)")
'''

pdf_func_idx = find_cell_by_content(nb, 'def export_to_pdf')
if pdf_func_idx:
    insert_cell_after(nb, pdf_func_idx, 'code', export_excel_code)
    print("  ✓ export_to_excel() 함수 추가 완료")

print("\nStep 4: export_report() 통합 함수 추가...")

export_report_code = '''def export_report(
    report_data: Dict[str, Any],
    format: str = 'pdf',
    output_dir: str = './reports'
) -> Dict[str, str]:
    """
    리포트를 지정된 형식으로 내보내기 (Phase 3)

    Args:
        report_data: 리포트 데이터
        format: 'pdf', 'excel', 'markdown', 'all' 중 하나
        output_dir: 저장 디렉토리

    Returns:
        {format: file_path} 딕셔너리
    """
    import os

    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    results = {}

    if format in ['pdf', 'all']:
        pdf_path = os.path.join(output_dir, f'ALM_Report_{timestamp}.pdf')
        export_to_pdf(report_data, pdf_path)
        results['pdf'] = pdf_path

    if format in ['excel', 'all']:
        excel_path = os.path.join(output_dir, f'ALM_Report_{timestamp}.xlsx')
        export_to_excel(report_data, excel_path)
        results['excel'] = excel_path

    if format in ['markdown', 'all']:
        md_path = os.path.join(output_dir, f'ALM_Report_{timestamp}.md')
        export_to_markdown(report_data, md_path)
        results['markdown'] = md_path

    return results

print("✓ 통합 내보내기 함수 정의 완료! (Phase 3)")
'''

excel_func_idx = find_cell_by_content(nb, 'def export_to_excel')
if excel_func_idx:
    insert_cell_after(nb, excel_func_idx, 'code', export_report_code)
    print("  ✓ export_report() 함수 추가 완료")

print("\n✓ Phase 3 함수 추가 완료!")

print("\n" + "="*70)
print("Phase 3, 4: Cell 12 (도구) 업데이트")
print("="*70)

# Cell 12 도구 업데이트
tools_idx = find_cell_by_content(nb, 'tools = [')
if tools_idx:
    existing_source = ''.join(nb['cells'][tools_idx].get('source', []))

    # Phase 2 도구 모델 추가
    phase2_models = '''
# Phase 2: 시나리오 비교 및 추세 분석 도구
class CompareScenariosInput(BaseModel):
    scenario_list: str = Field(description="비교할 시나리오 번호들 (쉼표로 구분, 예: '1,2,3')")
    comparison_metrics: str = Field(default="", description="비교할 지표 (선택사항)")

class AnalyzeTrendsInput(BaseModel):
    metric_type: str = Field(description="'exchange_rate' 또는 'interest_rate'")
    currency_or_rate_cd: str = Field(default="", description="통화 코드 또는 금리 코드 (선택사항)")
    start_date: str = Field(default="", description="시작 날짜 YYYY-MM-DD (선택사항)")
    end_date: str = Field(default="", description="종료 날짜 YYYY-MM-DD (선택사항)")

# Phase 1/3: 리포트 생성 및 내보내기 도구
class GenerateReportInput(BaseModel):
    include_sections: str = Field(default="", description="포함할 섹션 (쉼표 구분, 선택사항)")
    scenario_no: str = Field(default="", description="시나리오 번호 (선택사항)")

class ExportReportInput(BaseModel):
    format: str = Field(default="pdf", description="'pdf', 'excel', 'markdown', 'all' 중 하나")
    output_dir: str = Field(default="./reports", description="저장 디렉토리")

'''

    # Phase 2 wrapper 함수 추가
    phase2_wrappers = '''
def _compare_scenarios(scenario_list: str, comparison_metrics: str = "") -> str:
    """여러 시나리오를 비교 분석합니다."""
    scenarios = [int(s.strip()) for s in scenario_list.split(',')]

    metrics = None
    if comparison_metrics:
        metrics = [m.strip() for m in comparison_metrics.split(',')]

    result = compare_scenarios(scenarios, metrics)

    output = f"✓ 시나리오 비교 완료\\n\\n{result['summary']}\\n\\n"

    for scenario_no in scenarios:
        key = f'scenario_{scenario_no}'
        if key in result['comparison_data']:
            output += f"\\n--- 시나리오 {scenario_no} 상세 ---\\n"
            data = result['comparison_data'][key]['data'][:5]
            for row in data:
                output += f"{row}\\n"

    return output

def _analyze_trends(metric_type: str, currency_or_rate_cd: str = "",
                   start_date: str = "", end_date: str = "") -> str:
    """시계열 추세를 분석합니다."""
    result = analyze_trends(
        metric_type=metric_type,
        currency_or_rate_cd=currency_or_rate_cd if currency_or_rate_cd else None,
        start_date=start_date if start_date else None,
        end_date=end_date if end_date else None
    )

    if 'error' in result:
        return f"오류: {result['error']}"

    stats = result['statistics']
    output = f"✓ 추세 분석 완료 ({result['metric_type']})\\n\\n"
    output += f"추세: {result['trend']}\\n\\n"
    output += f"통계:\\n"
    output += f"  - 데이터 포인트: {stats['count']}개\\n"
    output += f"  - 평균: {stats['mean']:.4f}\\n"
    output += f"  - 표준편차: {stats['std']:.4f}\\n"
    output += f"  - 범위: {stats['min']:.4f} ~ {stats['max']:.4f}\\n"
    output += f"  - 변화: {stats['first_value']:.4f} → {stats['last_value']:.4f} ({stats['change_pct']:.2f}%)\\n"

    if 'slope' in stats:
        output += f"  - 기울기: {stats['slope']:.6f}\\n"

    return output

def _generate_report(include_sections: str = "", scenario_no: str = "") -> str:
    """종합 ALM 분석 리포트를 생성합니다."""
    sections = None
    if include_sections:
        sections = [s.strip() for s in include_sections.split(',')]

    scenario = None
    if scenario_no:
        scenario = int(scenario_no)

    report = generate_comprehensive_report(sections, scenario)

    output = f"✓ {report['title']} 생성 완료\\n\\n"
    output += f"생성일시: {report['generated_at'].strftime('%Y-%m-%d %H:%M:%S')}\\n"
    output += f"섹션 수: {len(report['sections'])}개\\n\\n"

    for section_name, section_data in report['sections'].items():
        output += f"- {section_data['title']}: {section_data.get('summary', '')}\\n"

    # 자동으로 Markdown 저장
    import os
    os.makedirs('./reports', exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    md_path = f'./reports/ALM_Report_{timestamp}.md'
    export_to_markdown(report, md_path)
    output += f"\\n✓ Markdown 파일 저장됨: {md_path}"

    return output

# 전역 변수로 마지막 리포트 저장 (export 시 사용)
_last_report = None

def _generate_report_with_storage(include_sections: str = "", scenario_no: str = "") -> str:
    """리포트 생성 + 전역 저장"""
    global _last_report

    sections = None
    if include_sections:
        sections = [s.strip() for s in include_sections.split(',')]

    scenario = None
    if scenario_no:
        scenario = int(scenario_no)

    _last_report = generate_comprehensive_report(sections, scenario)

    output = f"✓ {_last_report['title']} 생성 완료\\n\\n"
    output += f"생성일시: {_last_report['generated_at'].strftime('%Y-%m-%d %H:%M:%S')}\\n"
    output += f"섹션 수: {len(_last_report['sections'])}개\\n\\n"

    for section_name, section_data in _last_report['sections'].items():
        output += f"- {section_data['title']}: {section_data.get('summary', '')}\\n"

    return output

def _export_report(format: str = "pdf", output_dir: str = "./reports") -> str:
    """마지막 생성된 리포트를 지정 형식으로 내보냅니다."""
    global _last_report

    if _last_report is None:
        return "오류: 먼저 리포트를 생성해주세요 (generate_comprehensive_report 도구 사용)"

    results = export_report(_last_report, format, output_dir)

    output = f"✓ 리포트 내보내기 완료\\n\\n"
    for fmt, path in results.items():
        output += f"- {fmt.upper()}: {path}\\n"

    return output

'''

    # tools 리스트에 새 도구 추가
    new_tools = '''
    StructuredTool.from_function(
        func=_compare_scenarios,
        name="compare_scenarios",
        description="여러 시나리오의 유동성 갭을 비교 분석합니다",
        args_schema=CompareScenariosInput
    ),
    StructuredTool.from_function(
        func=_analyze_trends,
        name="analyze_trends",
        description="환율 또는 금리의 시계열 추세를 분석합니다",
        args_schema=AnalyzeTrendsInput
    ),
    StructuredTool.from_function(
        func=_generate_report_with_storage,
        name="generate_comprehensive_report",
        description="ALM 종합 분석 리포트를 생성합니다",
        args_schema=GenerateReportInput
    ),
    StructuredTool.from_function(
        func=_export_report,
        name="export_report",
        description="생성된 리포트를 PDF/Excel/Markdown 형식으로 내보냅니다",
        args_schema=ExportReportInput
    ),
]

print(f"✓ 총 {len(tools)}개의 도구가 정의되었습니다 (Phase 2, 3 완료)")
'''

    # 기존 tools 리스트 마지막에 추가
    updated_source = existing_source.replace(
        'print(f"✓ 총 {len(tools)}개의 도구가 정의되었습니다")',
        new_tools
    )

    # Pydantic 모델을 tools 정의 전에 추가
    updated_source = updated_source.replace(
        '# 도구 wrapper 함수',
        phase2_models + '\n# 도구 wrapper 함수' + phase2_wrappers
    )

    replace_cell_source(nb, tools_idx, updated_source)
    print("  ✓ Cell 12: 4개 도구 추가 완료 (총 9-10개 도구)")

print("\n" + "="*70)
print("Phase 4: Cell 19 (프롬프트) 업데이트")
print("="*70)

# Cell 19 프롬프트 업데이트
prompt_idx = find_cell_by_content(nb, 'SYSTEM_PROMPT =')
if prompt_idx:
    existing_source = ''.join(nb['cells'][prompt_idx].get('source', []))

    # 도구 목록에 Phase 2, 3 도구 추가
    updated_prompt = existing_source.replace(
        '6. visualize_data - 쿼리 결과를 차트로 시각화',
        '''6. visualize_data - 쿼리 결과를 차트로 시각화
7. generate_comprehensive_report - ALM 종합 분석 리포트 생성
8. compare_scenarios - 여러 시나리오 비교 분석
9. analyze_trends - 시계열 추세 분석 (환율, 금리)
10. export_report - 리포트를 PDF/Excel/Markdown으로 내보내기'''
    )

    # 리포트 생성 지침 추가
    report_instructions = '''

리포트 생성 시:
- 종합 분석 리포트: generate_comprehensive_report 도구 사용
- 시나리오 비교: compare_scenarios 도구 사용
- 추세 분석: analyze_trends 도구 사용
- 데이터 시각화: visualize_data 도구로 차트 생성
- 내보내기: export_report 도구로 PDF/Excel/Markdown 생성
- 리포트는 자동으로 ./reports 디렉토리에 저장됩니다
'''

    updated_prompt = updated_prompt.replace(
        '- 한국어로 친절하게 답변하세요',
        '- 한국어로 친절하게 답변하세요' + report_instructions
    )

    replace_cell_source(nb, prompt_idx, updated_prompt)
    print("  ✓ Cell 19: 시스템 프롬프트 업데이트 완료")

# 저장
with open(NOTEBOOK_PATH, 'w') as f:
    json.dump(nb, f, ensure_ascii=False, indent=1)

print("\n" + "="*70)
print("✅ Phase 3, 4 구현 완료!")
print("="*70)
print("\n📝 완료된 작업:")
print("  ✓ Phase 3: export_to_markdown(), export_to_pdf(), export_to_excel(), export_report() 추가")
print("  ✓ Phase 3: Cell 12에 4개 도구 추가 (compare_scenarios, analyze_trends, generate_comprehensive_report, export_report)")
print("  ✓ Phase 4: Cell 19 프롬프트 업데이트 (10개 도구 설명 + 리포트 생성 지침)")
print("\n다음 단계:")
print("  1. Jupyter Notebook 열기")
print("  2. 필요한 패키지 설치: !pip install reportlab openpyxl Pillow")
print("  3. 모든 셀 실행 (Kernel -> Restart & Run All)")
print("  4. 테스트:")
print('     chat("ALM 종합 리포트를 만들어줘")')
print('     chat("리포트를 PDF로 내보내줘")')
print('     chat("시나리오 1, 2를 비교해줘")')
print('     chat("USD 환율 추세를 분석해줘")')
