"""
ALM 챗봇 비즈니스 로직 함수
"""
import sqlite3
import pandas as pd
import os
from datetime import datetime
import json
from typing import Dict, List, Any, Optional
import numpy as np

# PDF/Excel 내보내기를 위한 라이브러리 (선택적)
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("Warning: reportlab not installed. PDF export will not be available.")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel export will not be available.")

# ======================================================================
# 데이터베이스 설정
# ======================================================================

# 데이터베이스 경로
DB_PATH = 'simple.db'

def get_db_connection():
    """데이터베이스 연결 생성"""
    return sqlite3.connect(DB_PATH)

def get_table_info():
    """데이터베이스의 모든 테이블 정보 조회"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 테이블 목록 조회
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
    tables = cursor.fetchall()
    
    table_info = {}
    for table in tables:
        table_name = table[0]
        cursor.execute(f"PRAGMA table_info({table_name});")
        columns = cursor.fetchall()
        table_info[table_name] = [col[1] for col in columns]  # 컬럼명만 추출
    
    conn.close()
    return table_info

# 테이블 정보 확인
tables = get_table_info()
print("데이터베이스 테이블:")
for table_name, columns in tables.items():
    print(f"\n{table_name}: {len(columns)}개 컬럼")
    print(f"  주요 컬럼: {', '.join(columns[:5])}...")

# ======================================================================
# 스키마 설명 조회 (캐싱)
# ======================================================================

# 컬럼 설명 캐시
_column_descriptions_cache: Dict[str, str] = {}

def get_column_description(table_name: str, column_name: str) -> Optional[str]:
    """
    컬럼 설명 조회 (캐싱 포함)

    Args:
        table_name: 테이블명 (예: 'ALM_INST', 'INST_ALM_01')
        column_name: 컬럼명 (예: 'DIM_PROD')

    Returns:
        컬럼 설명 (없으면 None)
    """
    cache_key = f"{table_name}.{column_name}"

    # 캐시 확인
    if cache_key in _column_descriptions_cache:
        return _column_descriptions_cache[cache_key]

    # DB 조회
    try:
        query = f"""
        SELECT description
        FROM column_descriptions
        WHERE table_name = '{table_name}'
          AND column_name = '{column_name}'
        LIMIT 1
        """

        result = execute_sql_query(query)

        if result['success'] and result['row_count'] > 0:
            desc = result['data'][0]['description']
            _column_descriptions_cache[cache_key] = desc
            return desc
        else:
            # 없으면 None 캐싱 (재조회 방지)
            _column_descriptions_cache[cache_key] = None
            return None
    except Exception as e:
        return None

def get_column_label(column_name: str, table_name: str = 'ALM_INST') -> str:
    """
    컬럼 레이블 반환 (형식: "컬럼명 (설명)")

    Args:
        column_name: 컬럼명
        table_name: 테이블명 (기본값: 'ALM_INST')

    Returns:
        "DIM_PROD (차원-상품코드)" 또는 "DIM_PROD" (설명 없을 시)
    """
    # ALM_INST에 설명이 없으면 INST_ALM_01 확인
    desc = get_column_description(table_name, column_name)
    if desc is None and table_name == 'ALM_INST':
        desc = get_column_description('INST_ALM_01', column_name)

    if desc:
        return f"{column_name} ({desc})"
    else:
        return column_name

# ======================================================================
# 비즈니스 로직 함수들
# ======================================================================

def execute_sql_query(query: str) -> Dict[str, Any]:
    """
    SQL 쿼리를 실행하고 결과를 반환
    
    Args:
        query: 실행할 SQL 쿼리
    
    Returns:
        결과 딕셔너리 (data, columns, row_count 포함)
    """
    try:
        conn = get_db_connection()
        df = pd.read_sql_query(query, conn)
        conn.close()
        
        return {
            "success": True,
            "data": df.to_dict('records'),
            "columns": df.columns.tolist(),
            "row_count": len(df),
            "dataframe": df
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "data": [],
            "columns": [],
            "row_count": 0
        }

def search_alm_contracts(filters: Optional[Dict] = None, limit: int = 10) -> str:
    """
    ALM 계약 정보 검색
    
    Args:
        filters: 필터 조건 (예: {"CURRENCY_CD": "KRW"})
        limit: 결과 제한 수
    
    Returns:
        검색 결과 문자열
    """
    query = "SELECT * FROM ALM_INST"
    
    if filters:
        conditions = [f"{k} = '{v}'" for k, v in filters.items()]
        query += " WHERE " + " AND ".join(conditions)
    
    query += f" LIMIT {limit}"
    
    result = execute_sql_query(query)
    
    if result["success"]:
        df = result["dataframe"]
        return f"검색 결과: {result['row_count']}건\n\n{df.to_string()}"
    else:
        return f"오류 발생: {result['error']}"

def analyze_liquidity_gap(scenario_no: Optional[int] = None) -> str:
    """
    유동성 갭 분석
    
    Args:
        scenario_no: 시나리오 번호
    
    Returns:
        분석 결과 문자열
    """
    query = """
    SELECT 
        TIME_BAND,
        SUM(GAP_PRN_TOTAL) as 총_원금갭,
        SUM(GAP_INT_TOTAL) as 총_이자갭,
        COUNT(*) as 건수
    FROM NFAR_LIQ_GAP_310524
    """
    
    if scenario_no is not None:
        query += f" WHERE SCENARIO_NO = {scenario_no}"
    
    query += " GROUP BY TIME_BAND ORDER BY TIME_BAND"
    
    result = execute_sql_query(query)
    
    if result["success"]:
        df = result["dataframe"]
        return f"유동성 갭 분석 결과:\n\n{df.to_string()}\n\n총 {result['row_count']}개 기간대"
    else:
        return f"오류 발생: {result['error']}"

def get_exchange_rate(currency: str, date: Optional[str] = None) -> str:
    """
    환율 정보 조회
    
    Args:
        currency: 통화 코드 (예: USD, EUR)
        date: 조회 날짜 (YYYY-MM-DD 형식)
    
    Returns:
        환율 정보 문자열
    """
    query = f"SELECT * FROM NFA_EXCH_RATE_HIST WHERE UNIT_CURRENCY_CD = '{currency}'"
    
    if date:
        query += f" AND EFFECTIVE_DATE = '{date}'"
    
    query += " ORDER BY EFFECTIVE_DATE DESC LIMIT 10"
    
    result = execute_sql_query(query)
    
    if result["success"]:
        df = result["dataframe"]
        return f"{currency} 환율 정보:\n\n{df.to_string()}"
    else:
        return f"오류 발생: {result['error']}"

def get_interest_rate(rate_cd: int, term: Optional[int] = None) -> str:
    """
    금리 정보 조회
    
    Args:
        rate_cd: 금리 코드
        term: 금리 기간
    
    Returns:
        금리 정보 문자열
    """
    query = f"SELECT * FROM NFA_IRC_RATE_HIST WHERE INT_RATE_CD = {rate_cd}"
    
    if term is not None:
        query += f" AND INT_RATE_TERM = {term}"
    
    query += " ORDER BY EFFECTIVE_DATE DESC LIMIT 10"
    
    result = execute_sql_query(query)
    
    if result["success"]:
        df = result["dataframe"]
        return f"금리 정보:\n\n{df.to_string()}"
    else:
        return f"오류 발생: {result['error']}"

def get_aggregate_stats(table_name: str, group_by: str, aggregate_col: str) -> str:
    """
    집계 통계 조회
    
    Args:
        table_name: 테이블명
        group_by: 그룹화 컬럼
        aggregate_col: 집계 컬럼
    
    Returns:
        통계 결과 문자열
    """
    query = f"""
    SELECT 
        {group_by},
        COUNT(*) as 건수,
        SUM({aggregate_col}) as 합계,
        AVG({aggregate_col}) as 평균,
        MIN({aggregate_col}) as 최소,
        MAX({aggregate_col}) as 최대
    FROM {table_name}
    GROUP BY {group_by}
    ORDER BY 합계 DESC
    LIMIT 20
    """
    
    result = execute_sql_query(query)
    
    if result["success"]:
        df = result["dataframe"]
        return f"{table_name} 테이블 {group_by}별 {aggregate_col} 집계:\n\n{df.to_string()}"
    else:
        return f"오류 발생: {result['error']}"




# ============================================================
# Phase 1: 종합 리포트 생성 함수
# ============================================================

def generate_comprehensive_report(
    include_sections: Optional[List[str]] = None,
    scenario_no: Optional[int] = None
) -> Dict[str, Any]:
    """
    종합 ALM 분석 리포트 생성 (Phase 1)

    Args:
        include_sections: 포함할 섹션 리스트 (None이면 모든 섹션)
        scenario_no: 유동성 갭 분석에 사용할 시나리오 번호

    Returns:
        리포트 데이터 딕셔너리
    """
    report = {
        'title': 'ALM 종합 분석 리포트',
        'generated_at': datetime.now(),
        'sections': {},
        'metadata': {
            'scenario_no': scenario_no,
            'requested_sections': include_sections
        }
    }

    all_sections = ['data_overview', 'liquidity_gap', 'market_data', 'dimensional_analysis']
    sections_to_include = include_sections if include_sections else all_sections

    # 1. Data Overview
    if 'data_overview' in sections_to_include:
        query = """
        SELECT
            CURRENCY_CD as 통화,
            COUNT(*) as 계약수,
            SUM(CUR_PAR_BAL) as 총잔액,
            AVG(INT_RATE) as 평균금리
        FROM ALM_INST
        GROUP BY CURRENCY_CD
        ORDER BY 총잔액 DESC
        """
        result = execute_sql_query(query)

        if result["success"]:
            report['sections']['data_overview'] = {
                'title': '데이터 개요',
                'data': result['data'],
                'summary': f"총 {sum([r['계약수'] for r in result['data']])}건의 계약, "
                          f"{len(result['data'])}개 통화"
            }

    # 2. Liquidity Gap
    if 'liquidity_gap' in sections_to_include:
        query = """
        SELECT
            TIME_BAND as 기간대,
            SUM(GAP_PRN_TOTAL) as 원금갭,
            SUM(GAP_INT_TOTAL) as 이자갭,
            SUM(GAP_PRN_TOTAL + GAP_INT_TOTAL) as 총갭
        FROM NFAR_LIQ_GAP_310524
        """

        if scenario_no is not None:
            query += f" WHERE SCENARIO_NO = {scenario_no}"

        query += " GROUP BY TIME_BAND ORDER BY TIME_BAND"

        result = execute_sql_query(query)

        if result["success"]:
            df = result['dataframe']
            total_gap = df['총갭'].sum() if '총갭' in df.columns else 0

            report['sections']['liquidity_gap'] = {
                'title': '유동성 갭 분석',
                'data': result['data'],
                'summary': f"총 {result['row_count']}개 기간대, 총갭: {total_gap:,.0f}",
                'scenario_no': scenario_no
            }

    # 3. Market Data
    if 'market_data' in sections_to_include:
        exchange_query = """
        SELECT
            UNIT_CURRENCY_CD as 통화,
            EFFECTIVE_DATE as 일자,
            EXCH_RATE as 환율
        FROM NFA_EXCH_RATE_HIST
        WHERE UNIT_CURRENCY_CD IN ('USD', 'EUR', 'JPY', 'CNY')
        ORDER BY EFFECTIVE_DATE DESC
        LIMIT 20
        """

        exchange_result = execute_sql_query(exchange_query)

        interest_query = """
        SELECT
            INT_RATE_CD as 금리코드,
            INT_RATE_TERM as 기간,
            EFFECTIVE_DATE as 일자,
            INT_RATE as 금리
        FROM NFA_IRC_RATE_HIST
        ORDER BY EFFECTIVE_DATE DESC
        LIMIT 20
        """

        interest_result = execute_sql_query(interest_query)

        report['sections']['market_data'] = {
            'title': '시장 데이터',
            'exchange_rates': exchange_result['data'] if exchange_result['success'] else [],
            'interest_rates': interest_result['data'] if interest_result['success'] else [],
            'summary': f"환율 {len(exchange_result['data'])}건, 금리 {len(interest_result['data'])}건"
        }

    # 4. Dimensional Analysis
    if 'dimensional_analysis' in sections_to_include:
        dim_query = """
        SELECT
            ALM_DIMN_CD as ALM차원,
            COUNT(*) as 건수,
            SUM(CUR_PAR_BAL) as 총잔액
        FROM ALM_INST
        GROUP BY ALM_DIMN_CD
        ORDER BY 총잔액 DESC
        LIMIT 10
        """

        dim_result = execute_sql_query(dim_query)

        report['sections']['dimensional_analysis'] = {
            'title': '차원 분석',
            'data': dim_result['data'] if dim_result['success'] else [],
            'summary': f"총 {len(dim_result['data'])}개 ALM 차원"
        }

    return report


# ============================================================
# Phase 2: 시나리오 비교 함수
# ============================================================

def compare_scenarios(
    scenario_list: List[int],
    comparison_metrics: Optional[List[str]] = None
) -> Dict[str, Any]:
    """
    여러 시나리오 비교 분석 (Phase 2)

    Args:
        scenario_list: 비교할 시나리오 번호 리스트
        comparison_metrics: 비교할 지표 (None이면 기본 지표)

    Returns:
        {
            'scenarios': List[int],
            'comparison_data': Dict,
            'summary': str
        }
    """
    comparison = {
        'scenarios': scenario_list,
        'comparison_data': {},
        'summary': ''
    }

    for scenario_no in scenario_list:
        query = f"""
        SELECT
            TIME_BAND as 기간대,
            SUM(GAP_PRN_TOTAL) as 원금갭,
            SUM(GAP_INT_TOTAL) as 이자갭,
            SUM(GAP_PRN_TOTAL + GAP_INT_TOTAL) as 총갭
        FROM NFAR_LIQ_GAP_310524
        WHERE SCENARIO_NO = {scenario_no}
        GROUP BY TIME_BAND
        ORDER BY TIME_BAND
        """

        result = execute_sql_query(query)

        if result['success']:
            df = result['dataframe']

            comparison['comparison_data'][f'scenario_{scenario_no}'] = {
                'data': result['data'],
                'total_gap': df['총갭'].sum() if '총갭' in df.columns else 0,
                'max_gap': df['총갭'].max() if '총갭' in df.columns else 0,
                'min_gap': df['총갭'].min() if '총갭' in df.columns else 0,
                'avg_gap': df['총갭'].mean() if '총갭' in df.columns else 0
            }

    summary_lines = []
    summary_lines.append(f"총 {len(scenario_list)}개 시나리오 비교\n")

    for scenario_no in scenario_list:
        key = f'scenario_{scenario_no}'
        if key in comparison['comparison_data']:
            data = comparison['comparison_data'][key]
            summary_lines.append(
                f"시나리오 {scenario_no}: "
                f"총갭={data['total_gap']:,.0f}, "
                f"평균={data['avg_gap']:,.0f}, "
                f"최대={data['max_gap']:,.0f}, "
                f"최소={data['min_gap']:,.0f}"
            )

    comparison['summary'] = '\n'.join(summary_lines)
    return comparison


# ============================================================
# Phase 2: 추세 분석 함수
# ============================================================

def analyze_trends(
    metric_type: str,
    currency_or_rate_cd: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
) -> Dict[str, Any]:
    """
    시계열 추세 분석 (Phase 2)

    Args:
        metric_type: 'exchange_rate' 또는 'interest_rate'
        currency_or_rate_cd: 통화 코드 또는 금리 코드
        start_date: 시작 날짜 (YYYY-MM-DD)
        end_date: 종료 날짜 (YYYY-MM-DD)

    Returns:
        {
            'metric_type': str,
            'data_points': List[Dict],
            'statistics': Dict,
            'trend': str
        }
    """
    import numpy as np

    trends = {
        'metric_type': metric_type,
        'data_points': [],
        'statistics': {},
        'trend': ''
    }

    if metric_type == 'exchange_rate':
        query = "SELECT EFFECTIVE_DATE as 일자, EXCH_RATE as 값 FROM NFA_EXCH_RATE_HIST"

        conditions = []
        if currency_or_rate_cd:
            conditions.append(f"UNIT_CURRENCY_CD = '{currency_or_rate_cd}'")
        if start_date:
            conditions.append(f"EFFECTIVE_DATE >= '{start_date}'")
        if end_date:
            conditions.append(f"EFFECTIVE_DATE <= '{end_date}'")

        if conditions:
            query += " WHERE " + " AND ".join(conditions)

        query += " ORDER BY EFFECTIVE_DATE"

    elif metric_type == 'interest_rate':
        query = "SELECT EFFECTIVE_DATE as 일자, INT_RATE as 값 FROM NFA_IRC_RATE_HIST"

        conditions = []
        if currency_or_rate_cd:
            conditions.append(f"INT_RATE_CD = {currency_or_rate_cd}")
        if start_date:
            conditions.append(f"EFFECTIVE_DATE >= '{start_date}'")
        if end_date:
            conditions.append(f"EFFECTIVE_DATE <= '{end_date}'")

        if conditions:
            query += " WHERE " + " AND ".join(conditions)

        query += " ORDER BY EFFECTIVE_DATE"

    else:
        return {
            'error': f"지원하지 않는 metric_type: {metric_type}"
        }

    result = execute_sql_query(query)

    if result['success'] and result['row_count'] > 0:
        df = result['dataframe']
        trends['data_points'] = result['data']

        values = df['값'].values
        trends['statistics'] = {
            'count': len(values),
            'mean': float(np.mean(values)),
            'std': float(np.std(values)),
            'min': float(np.min(values)),
            'max': float(np.max(values)),
            'first_value': float(values[0]),
            'last_value': float(values[-1]),
            'change': float(values[-1] - values[0]),
            'change_pct': float((values[-1] - values[0]) / values[0] * 100) if values[0] != 0 else 0
        }

        if len(values) >= 2:
            x = np.arange(len(values))
            slope = np.polyfit(x, values, 1)[0]

            if slope > 0.01:
                trends['trend'] = '상승 추세'
            elif slope < -0.01:
                trends['trend'] = '하락 추세'
            else:
                trends['trend'] = '안정 추세'

            trends['statistics']['slope'] = float(slope)

    return trends


# ============================================================
# 신규 포지션 증가분 분석
# ============================================================

def analyze_new_position_growth(
    current_base_date: str,
    previous_base_date: Optional[str] = None,
    group_by_dimensions: Optional[List[str]] = None
) -> Dict[str, Any]:
    """
    당월 신규 포지션 증가분 분석

    이전 기준일 대비 새로 추가된 계약(REFERENCE_NO)을 식별하고
    차원별로 집계하여 신규 포지션 증가분을 분석합니다.

    Args:
        current_base_date: 현재 기준일 (YYYY-MM-DD)
        previous_base_date: 이전 기준일 (None이면 자동으로 직전 BASE_DATE 선택)
        group_by_dimensions: 그룹화 차원 리스트 ['DIM_PROD', 'DIM_ORG', 'DIM_ALM']
                            None이면 모든 차원으로 분석

    Returns:
        {
            'current_date': str,
            'previous_date': str,
            'new_contracts': {
                'count': int,
                'total_balance': float,
                'contracts': List[Dict]  # 신규 계약 샘플 (최대 5건)
            },
            'dimensional_breakdown': {
                'by_product': List[Dict],   # DIM_PROD별 신규 집계
                'by_org': List[Dict],       # DIM_ORG별 신규 집계
                'by_alm': List[Dict]        # DIM_ALM별 신규 집계
            },
            'summary': str
        }
    """

    # 1. previous_base_date 자동 선택 (None인 경우)
    if previous_base_date is None:
        query = f"""
        SELECT DISTINCT BASE_DATE
        FROM ALM_INST
        WHERE BASE_DATE LIKE '{current_base_date}%'
           OR BASE_DATE < '{current_base_date}'
        ORDER BY BASE_DATE DESC
        LIMIT 2
        """

        result = execute_sql_query(query)

        if result['success'] and result['row_count'] >= 2:
            # Get the second one (first is current_base_date itself)
            previous_base_date = result['data'][1]['BASE_DATE']
        elif result['success'] and result['row_count'] == 1:
            # Only one date found, try to get previous
            query = f"""
            SELECT DISTINCT BASE_DATE
            FROM ALM_INST
            WHERE BASE_DATE < '{current_base_date}'
            ORDER BY BASE_DATE DESC
            LIMIT 1
            """
            result2 = execute_sql_query(query)
            if result2['success'] and result2['row_count'] > 0:
                previous_base_date = result2['data'][0]['BASE_DATE']
            else:
                return {
                    'error': f"이전 기준일을 찾을 수 없습니다. {current_base_date}보다 이전 데이터가 없습니다."
                }
        else:
            return {
                'error': f"이전 기준일을 찾을 수 없습니다. {current_base_date}보다 이전 데이터가 없습니다."
            }

    # 2. 신규 계약 식별 및 기본 통계
    new_contracts_query = f"""
    SELECT
        t1.REFERENCE_NO,
        t1.CURRENCY_CD,
        t1.CUR_PAR_BAL,
        t1.CUR_RATE,
        t1.DIM_PROD,
        t1.DIM_ORG,
        t1.DIM_ALM
    FROM ALM_INST t1
    LEFT JOIN ALM_INST t2
        ON t1.REFERENCE_NO = t2.REFERENCE_NO
        AND t2.BASE_DATE LIKE '{previous_base_date[:10]}%'
    WHERE t1.BASE_DATE LIKE '{current_base_date[:10]}%'
      AND t2.REFERENCE_NO IS NULL
    """

    new_contracts_result = execute_sql_query(new_contracts_query)

    if not new_contracts_result['success']:
        return {
            'error': f"신규 계약 조회 실패: {new_contracts_result['error']}"
        }

    df = new_contracts_result['dataframe']

    # 신규 계약이 없는 경우
    if len(df) == 0:
        return {
            'current_date': current_base_date,
            'previous_date': previous_base_date,
            'new_contracts': {
                'count': 0,
                'total_balance': 0,
                'contracts': []
            },
            'dimensional_breakdown': {
                'by_product': [],
                'by_org': [],
                'by_alm': []
            },
            'summary': f"{current_base_date} 기준 신규 계약이 없습니다 (비교: {previous_base_date})"
        }

    # 기본 통계
    new_count = len(df)
    total_balance = df['CUR_PAR_BAL'].sum() if 'CUR_PAR_BAL' in df.columns else 0

    # 샘플 계약 (최대 5건)
    sample_contracts = new_contracts_result['data'][:5]

    # 3. 차원별 집계
    dimensional_breakdown = {
        'by_product': [],
        'by_org': [],
        'by_alm': []
    }

    # 기본 차원 설정
    if group_by_dimensions is None:
        group_by_dimensions = ['DIM_PROD', 'DIM_ORG', 'DIM_ALM']

    # DIM_PROD별 집계
    if 'DIM_PROD' in group_by_dimensions:
        dim_query = f"""
        SELECT
            t1.DIM_PROD as 차원값,
            COUNT(DISTINCT t1.REFERENCE_NO) as 신규건수,
            SUM(t1.CUR_PAR_BAL) as 신규잔액
        FROM ALM_INST t1
        LEFT JOIN ALM_INST t2
            ON t1.REFERENCE_NO = t2.REFERENCE_NO
            AND t2.BASE_DATE LIKE '{previous_base_date[:10]}%'
        WHERE t1.BASE_DATE LIKE '{current_base_date[:10]}%'
          AND t2.REFERENCE_NO IS NULL
          AND t1.DIM_PROD IS NOT NULL
        GROUP BY t1.DIM_PROD
        ORDER BY 신규잔액 DESC
        """

        dim_result = execute_sql_query(dim_query)
        if dim_result['success']:
            dimensional_breakdown['by_product'] = dim_result['data']

    # DIM_ORG별 집계
    if 'DIM_ORG' in group_by_dimensions:
        dim_query = f"""
        SELECT
            t1.DIM_ORG as 차원값,
            COUNT(DISTINCT t1.REFERENCE_NO) as 신규건수,
            SUM(t1.CUR_PAR_BAL) as 신규잔액
        FROM ALM_INST t1
        LEFT JOIN ALM_INST t2
            ON t1.REFERENCE_NO = t2.REFERENCE_NO
            AND t2.BASE_DATE LIKE '{previous_base_date[:10]}%'
        WHERE t1.BASE_DATE LIKE '{current_base_date[:10]}%'
          AND t2.REFERENCE_NO IS NULL
          AND t1.DIM_ORG IS NOT NULL
        GROUP BY t1.DIM_ORG
        ORDER BY 신규잔액 DESC
        """

        dim_result = execute_sql_query(dim_query)
        if dim_result['success']:
            dimensional_breakdown['by_org'] = dim_result['data']

    # DIM_ALM별 집계
    if 'DIM_ALM' in group_by_dimensions:
        dim_query = f"""
        SELECT
            t1.DIM_ALM as 차원값,
            COUNT(DISTINCT t1.REFERENCE_NO) as 신규건수,
            SUM(t1.CUR_PAR_BAL) as 신규잔액
        FROM ALM_INST t1
        LEFT JOIN ALM_INST t2
            ON t1.REFERENCE_NO = t2.REFERENCE_NO
            AND t2.BASE_DATE LIKE '{previous_base_date[:10]}%'
        WHERE t1.BASE_DATE LIKE '{current_base_date[:10]}%'
          AND t2.REFERENCE_NO IS NULL
          AND t1.DIM_ALM IS NOT NULL
        GROUP BY t1.DIM_ALM
        ORDER BY 신규잔액 DESC
        """

        dim_result = execute_sql_query(dim_query)
        if dim_result['success']:
            dimensional_breakdown['by_alm'] = dim_result['data']

    # 4. 요약 생성
    summary = f"{current_base_date} 기준 신규 {new_count}건, 총 잔액 {total_balance:,.0f} (비교: {previous_base_date})"

    return {
        'current_date': current_base_date,
        'previous_date': previous_base_date,
        'new_contracts': {
            'count': new_count,
            'total_balance': float(total_balance),
            'contracts': sample_contracts
        },
        'dimensional_breakdown': dimensional_breakdown,
        'summary': summary
    }


# ============================================================
# 소멸 포지션 감소분 분석
# ============================================================

def analyze_expired_position_decrease(
    current_base_date: str,
    previous_base_date: Optional[str] = None,
    group_by_dimensions: Optional[List[str]] = None
) -> Dict[str, Any]:
    """
    당월 소멸 포지션 감소분 분석

    이전 기준일에는 존재했지만 현재 기준일에는 사라진 계약(REFERENCE_NO)을 식별하고
    차원별로 집계하여 소멸 포지션 감소분을 분석합니다.

    Args:
        current_base_date: 현재 기준일 (YYYY-MM-DD)
        previous_base_date: 이전 기준일 (None이면 자동으로 직전 BASE_DATE 선택)
        group_by_dimensions: 그룹화 차원 리스트 ['DIM_PROD', 'DIM_ORG', 'DIM_ALM']
                            None이면 모든 차원으로 분석

    Returns:
        {
            'current_date': str,
            'previous_date': str,
            'expired_contracts': {
                'count': int,
                'total_balance': float,
                'contracts': List[Dict]  # 소멸 계약 샘플 (최대 5건)
            },
            'dimensional_breakdown': {
                'by_product': List[Dict],   # DIM_PROD별 소멸 집계
                'by_org': List[Dict],       # DIM_ORG별 소멸 집계
                'by_alm': List[Dict]        # DIM_ALM별 소멸 집계
            },
            'summary': str
        }
    """

    # 1. previous_base_date 자동 선택 (None인 경우)
    if previous_base_date is None:
        query = f"""
        SELECT DISTINCT BASE_DATE
        FROM ALM_INST
        WHERE BASE_DATE LIKE '{current_base_date}%'
           OR BASE_DATE < '{current_base_date}'
        ORDER BY BASE_DATE DESC
        LIMIT 2
        """

        result = execute_sql_query(query)

        if result['success'] and result['row_count'] >= 2:
            previous_base_date = result['data'][1]['BASE_DATE']
        elif result['success'] and result['row_count'] == 1:
            query = f"""
            SELECT DISTINCT BASE_DATE
            FROM ALM_INST
            WHERE BASE_DATE < '{current_base_date}'
            ORDER BY BASE_DATE DESC
            LIMIT 1
            """
            result2 = execute_sql_query(query)
            if result2['success'] and result2['row_count'] > 0:
                previous_base_date = result2['data'][0]['BASE_DATE']
            else:
                return {
                    'error': f"이전 기준일을 찾을 수 없습니다. {current_base_date}보다 이전 데이터가 없습니다."
                }
        else:
            return {
                'error': f"이전 기준일을 찾을 수 없습니다. {current_base_date}보다 이전 데이터가 없습니다."
            }

    # 2. 소멸 계약 식별 (이전에는 있었지만 현재는 없는 계약)
    expired_contracts_query = f"""
    SELECT
        t1.REFERENCE_NO,
        t1.CURRENCY_CD,
        t1.CUR_PAR_BAL,
        t1.CUR_RATE,
        t1.DIM_PROD,
        t1.DIM_ORG,
        t1.DIM_ALM
    FROM ALM_INST t1
    LEFT JOIN ALM_INST t2
        ON t1.REFERENCE_NO = t2.REFERENCE_NO
        AND t2.BASE_DATE LIKE '{current_base_date[:10]}%'
    WHERE t1.BASE_DATE LIKE '{previous_base_date[:10]}%'
      AND t2.REFERENCE_NO IS NULL
    """

    expired_contracts_result = execute_sql_query(expired_contracts_query)

    if not expired_contracts_result['success']:
        return {
            'error': f"소멸 계약 조회 실패: {expired_contracts_result['error']}"
        }

    df = expired_contracts_result['dataframe']

    # 소멸 계약이 없는 경우
    if len(df) == 0:
        return {
            'current_date': current_base_date,
            'previous_date': previous_base_date,
            'expired_contracts': {
                'count': 0,
                'total_balance': 0,
                'contracts': []
            },
            'dimensional_breakdown': {
                'by_product': [],
                'by_org': [],
                'by_alm': []
            },
            'summary': f"{current_base_date} 기준 소멸 계약이 없습니다 (비교: {previous_base_date})"
        }

    # 기본 통계
    expired_count = len(df)
    total_balance = df['CUR_PAR_BAL'].sum() if 'CUR_PAR_BAL' in df.columns else 0

    # 샘플 계약 (최대 5건)
    sample_contracts = expired_contracts_result['data'][:5]

    # 3. 차원별 집계
    dimensional_breakdown = {
        'by_product': [],
        'by_org': [],
        'by_alm': []
    }

    # 기본 차원 설정
    if group_by_dimensions is None:
        group_by_dimensions = ['DIM_PROD', 'DIM_ORG', 'DIM_ALM']

    # DIM_PROD별 집계
    if 'DIM_PROD' in group_by_dimensions:
        dim_query = f"""
        SELECT
            t1.DIM_PROD as 차원값,
            COUNT(DISTINCT t1.REFERENCE_NO) as 소멸건수,
            SUM(t1.CUR_PAR_BAL) as 소멸잔액
        FROM ALM_INST t1
        LEFT JOIN ALM_INST t2
            ON t1.REFERENCE_NO = t2.REFERENCE_NO
            AND t2.BASE_DATE LIKE '{current_base_date[:10]}%'
        WHERE t1.BASE_DATE LIKE '{previous_base_date[:10]}%'
          AND t2.REFERENCE_NO IS NULL
          AND t1.DIM_PROD IS NOT NULL
        GROUP BY t1.DIM_PROD
        ORDER BY 소멸잔액 DESC
        """

        dim_result = execute_sql_query(dim_query)
        if dim_result['success']:
            dimensional_breakdown['by_product'] = dim_result['data']

    # DIM_ORG별 집계
    if 'DIM_ORG' in group_by_dimensions:
        dim_query = f"""
        SELECT
            t1.DIM_ORG as 차원값,
            COUNT(DISTINCT t1.REFERENCE_NO) as 소멸건수,
            SUM(t1.CUR_PAR_BAL) as 소멸잔액
        FROM ALM_INST t1
        LEFT JOIN ALM_INST t2
            ON t1.REFERENCE_NO = t2.REFERENCE_NO
            AND t2.BASE_DATE LIKE '{current_base_date[:10]}%'
        WHERE t1.BASE_DATE LIKE '{previous_base_date[:10]}%'
          AND t2.REFERENCE_NO IS NULL
          AND t1.DIM_ORG IS NOT NULL
        GROUP BY t1.DIM_ORG
        ORDER BY 소멸잔액 DESC
        """

        dim_result = execute_sql_query(dim_query)
        if dim_result['success']:
            dimensional_breakdown['by_org'] = dim_result['data']

    # DIM_ALM별 집계
    if 'DIM_ALM' in group_by_dimensions:
        dim_query = f"""
        SELECT
            t1.DIM_ALM as 차원값,
            COUNT(DISTINCT t1.REFERENCE_NO) as 소멸건수,
            SUM(t1.CUR_PAR_BAL) as 소멸잔액
        FROM ALM_INST t1
        LEFT JOIN ALM_INST t2
            ON t1.REFERENCE_NO = t2.REFERENCE_NO
            AND t2.BASE_DATE LIKE '{current_base_date[:10]}%'
        WHERE t1.BASE_DATE LIKE '{previous_base_date[:10]}%'
          AND t2.REFERENCE_NO IS NULL
          AND t1.DIM_ALM IS NOT NULL
        GROUP BY t1.DIM_ALM
        ORDER BY 소멸잔액 DESC
        """

        dim_result = execute_sql_query(dim_query)
        if dim_result['success']:
            dimensional_breakdown['by_alm'] = dim_result['data']

    # 4. 요약 생성
    summary = f"{current_base_date} 기준 소멸 {expired_count}건, 총 잔액 {total_balance:,.0f} (비교: {previous_base_date})"

    return {
        'current_date': current_base_date,
        'previous_date': previous_base_date,
        'expired_contracts': {
            'count': expired_count,
            'total_balance': float(total_balance),
            'contracts': sample_contracts
        },
        'dimensional_breakdown': dimensional_breakdown,
        'summary': summary
    }


# ============================================================
# Phase 1,3: 내보내기 함수들
# ============================================================

def export_to_markdown(report_data: Dict[str, Any], output_path: str) -> str:
    """
    리포트를 Markdown 형식으로 내보내기 (Phase 1)
    """
    import os

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)

    lines = []

    # 제목
    lines.append(f"# {report_data['title']}\n")
    lines.append(f"생성일시: {report_data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')}\n")

    if report_data['metadata'].get('scenario_no'):
        lines.append(f"시나리오: {report_data['metadata']['scenario_no']}\n")

    lines.append("\n---\n\n")

    # 각 섹션
    sections = report_data.get('sections', {})

    for section_name, section_data in sections.items():
        lines.append(f"## {section_data['title']}\n\n")
        lines.append(f"{section_data.get('summary', '')}\n\n")

        # 데이터 테이블
        if 'data' in section_data and section_data['data']:
            data = section_data['data'][:20]
            if data:
                headers = list(data[0].keys())
                lines.append("| " + " | ".join(headers) + " |\n")
                lines.append("| " + " | ".join(['---'] * len(headers)) + " |\n")

                for row in data:
                    values = [str(row.get(h, '')) for h in headers]
                    lines.append("| " + " | ".join(values) + " |\n")

                lines.append("\n")

        lines.append("\n")

    with open(output_path, 'w', encoding='utf-8') as f:
        f.writelines(lines)

    return output_path


def export_to_pdf(report_data: Dict[str, Any], output_path: str) -> str:
    """
    리포트를 PDF 형식으로 내보내기 (Phase 3)
    """
    if not REPORTLAB_AVAILABLE:
        return "오류: reportlab 패키지가 설치되어 있지 않습니다. pip install reportlab을 실행하세요."

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)

    doc = SimpleDocTemplate(output_path, pagesize=A4)
    story = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30
    )
    heading_style = styles['Heading2']

    story.append(Paragraph(report_data['title'], title_style))
    story.append(Spacer(1, 0.5*cm))

    meta_text = f"생성일시: {report_data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')}"
    if report_data['metadata'].get('scenario_no'):
        meta_text += f" | 시나리오: {report_data['metadata']['scenario_no']}"
    story.append(Paragraph(meta_text, styles['Normal']))
    story.append(Spacer(1, 1*cm))

    sections = report_data.get('sections', {})

    for section_name, section_data in sections.items():
        story.append(Paragraph(section_data['title'], heading_style))
        story.append(Spacer(1, 0.3*cm))

        story.append(Paragraph(section_data.get('summary', ''), styles['Normal']))
        story.append(Spacer(1, 0.5*cm))

        if 'data' in section_data and section_data['data']:
            data = section_data['data'][:10]
            if data:
                headers = list(data[0].keys())
                table_data = [headers]

                for row in data:
                    table_data.append([str(row.get(h, '')) for h in headers])

                t = Table(table_data)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))

                story.append(t)
                story.append(Spacer(1, 0.5*cm))

        story.append(Spacer(1, 0.5*cm))

    doc.build(story)

    return output_path


def export_to_excel(report_data: Dict[str, Any], output_path: str) -> str:
    """
    리포트를 Excel 형식으로 내보내기 (Phase 3)
    """
    if not OPENPYXL_AVAILABLE:
        return "오류: openpyxl 패키지가 설치되어 있지 않습니다. pip install openpyxl을 실행하세요."

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)

    wb = Workbook()

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    summary_ws = wb.create_sheet(title="요약")
    summary_ws.append([report_data['title']])
    summary_ws.append([])
    summary_ws.append(['생성일시', report_data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')])

    if report_data['metadata'].get('scenario_no'):
        summary_ws.append(['시나리오', report_data['metadata']['scenario_no']])

    summary_ws.append([])
    summary_ws.append(['섹션', '요약'])

    title_font = Font(size=16, bold=True)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    summary_ws['A1'].font = title_font

    sections = report_data.get('sections', {})

    for section_name, section_data in sections.items():
        summary_ws.append([section_data['title'], section_data.get('summary', '')])

        ws = wb.create_sheet(title=section_data['title'][:30])

        ws.append([section_data['title']])
        ws.append([section_data.get('summary', '')])
        ws.append([])

        if 'data' in section_data and section_data['data']:
            data = section_data['data']
            if data:
                headers = list(data[0].keys())
                ws.append(headers)

                for col_idx, _ in enumerate(headers, 1):
                    cell = ws.cell(row=ws.max_row, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')

                for row_data in data:
                    ws.append([row_data.get(h, '') for h in headers])

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)

    return output_path


def export_report(
    report_data: Dict[str, Any],
    format: str = 'pdf',
    output_dir: str = './reports'
) -> Dict[str, str]:
    """
    리포트를 지정된 형식으로 내보내기 (Phase 3)
    """
    import os

    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    results = {}

    if format in ['pdf', 'all']:
        pdf_path = os.path.join(output_dir, f'ALM_Report_{timestamp}.pdf')
        export_to_pdf(report_data, pdf_path)
        results['pdf'] = pdf_path

    if format in ['excel', 'all']:
        excel_path = os.path.join(output_dir, f'ALM_Report_{timestamp}.xlsx')
        export_to_excel(report_data, excel_path)
        results['excel'] = excel_path

    if format in ['markdown', 'all']:
        md_path = os.path.join(output_dir, f'ALM_Report_{timestamp}.md')
        export_to_markdown(report_data, md_path)
        results['markdown'] = md_path

    return results

print("✓ Phase 1-3 모든 리포트 함수 정의 완료!")
