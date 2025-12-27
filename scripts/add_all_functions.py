#!/usr/bin/env python3
"""
Phase 1-3 모든 리포트 함수 추가 스크립트
Cell 9 끝에 모든 함수를 한꺼번에 추가
"""

import json

NOTEBOOK_PATH = "chatbot.ipynb"

print("="*70)
print("Phase 1-3 모든 리포트 함수 추가")
print("="*70)

# 노트북 로드
with open(NOTEBOOK_PATH, 'r') as f:
    nb = json.load(f)

print(f"\n✓ 현재 노트북: {len(nb['cells'])}개 셀")

# Cell 9 찾기 (execute_sql_query가 있는 셀)
cell9_idx = None
for idx, cell in enumerate(nb['cells']):
    if cell['cell_type'] == 'code':
        source = ''.join(cell.get('source', []))
        if 'def execute_sql_query' in source:
            cell9_idx = idx
            break

if cell9_idx is None:
    print("ERROR: Cell 9를 찾을 수 없습니다!")
    exit(1)

print(f"✓ Cell {cell9_idx} 발견")

# 현재 Cell 9 소스 가져오기
current_source = ''.join(nb['cells'][cell9_idx].get('source', []))

# 추가할 모든 함수들
all_new_functions = '''

# ============================================================
# Phase 1: 종합 리포트 생성 함수
# ============================================================

def generate_comprehensive_report(
    include_sections: Optional[List[str]] = None,
    scenario_no: Optional[int] = None
) -> Dict[str, Any]:
    """
    종합 ALM 분석 리포트 생성 (Phase 1)

    Args:
        include_sections: 포함할 섹션 리스트 (None이면 모든 섹션)
        scenario_no: 유동성 갭 분석에 사용할 시나리오 번호

    Returns:
        리포트 데이터 딕셔너리
    """
    report = {
        'title': 'ALM 종합 분석 리포트',
        'generated_at': datetime.now(),
        'sections': {},
        'metadata': {
            'scenario_no': scenario_no,
            'requested_sections': include_sections
        }
    }

    all_sections = ['data_overview', 'liquidity_gap', 'market_data', 'dimensional_analysis']
    sections_to_include = include_sections if include_sections else all_sections

    # 1. Data Overview
    if 'data_overview' in sections_to_include:
        query = """
        SELECT
            CURRENCY_CD as 통화,
            COUNT(*) as 계약수,
            SUM(CUR_PAR_BAL) as 총잔액,
            AVG(INT_RATE) as 평균금리
        FROM ALM_INST
        GROUP BY CURRENCY_CD
        ORDER BY 총잔액 DESC
        """
        result = execute_sql_query(query)

        if result["success"]:
            report['sections']['data_overview'] = {
                'title': '데이터 개요',
                'data': result['data'],
                'summary': f"총 {sum([r['계약수'] for r in result['data']])}건의 계약, "
                          f"{len(result['data'])}개 통화"
            }

    # 2. Liquidity Gap
    if 'liquidity_gap' in sections_to_include:
        query = """
        SELECT
            TIME_BAND as 기간대,
            SUM(GAP_PRN_TOTAL) as 원금갭,
            SUM(GAP_INT_TOTAL) as 이자갭,
            SUM(GAP_PRN_TOTAL + GAP_INT_TOTAL) as 총갭
        FROM NFAR_LIQ_GAP_310524
        """

        if scenario_no is not None:
            query += f" WHERE SCENARIO_NO = {scenario_no}"

        query += " GROUP BY TIME_BAND ORDER BY TIME_BAND"

        result = execute_sql_query(query)

        if result["success"]:
            df = result['dataframe']
            total_gap = df['총갭'].sum() if '총갭' in df.columns else 0

            report['sections']['liquidity_gap'] = {
                'title': '유동성 갭 분석',
                'data': result['data'],
                'summary': f"총 {result['row_count']}개 기간대, 총갭: {total_gap:,.0f}",
                'scenario_no': scenario_no
            }

    # 3. Market Data
    if 'market_data' in sections_to_include:
        exchange_query = """
        SELECT
            UNIT_CURRENCY_CD as 통화,
            EFFECTIVE_DATE as 일자,
            EXCH_RATE as 환율
        FROM NFA_EXCH_RATE_HIST
        WHERE UNIT_CURRENCY_CD IN ('USD', 'EUR', 'JPY', 'CNY')
        ORDER BY EFFECTIVE_DATE DESC
        LIMIT 20
        """

        exchange_result = execute_sql_query(exchange_query)

        interest_query = """
        SELECT
            INT_RATE_CD as 금리코드,
            INT_RATE_TERM as 기간,
            EFFECTIVE_DATE as 일자,
            INT_RATE as 금리
        FROM NFA_IRC_RATE_HIST
        ORDER BY EFFECTIVE_DATE DESC
        LIMIT 20
        """

        interest_result = execute_sql_query(interest_query)

        report['sections']['market_data'] = {
            'title': '시장 데이터',
            'exchange_rates': exchange_result['data'] if exchange_result['success'] else [],
            'interest_rates': interest_result['data'] if interest_result['success'] else [],
            'summary': f"환율 {len(exchange_result['data'])}건, 금리 {len(interest_result['data'])}건"
        }

    # 4. Dimensional Analysis
    if 'dimensional_analysis' in sections_to_include:
        dim_query = """
        SELECT
            ALM_DIMN_CD as ALM차원,
            COUNT(*) as 건수,
            SUM(CUR_PAR_BAL) as 총잔액
        FROM ALM_INST
        GROUP BY ALM_DIMN_CD
        ORDER BY 총잔액 DESC
        LIMIT 10
        """

        dim_result = execute_sql_query(dim_query)

        report['sections']['dimensional_analysis'] = {
            'title': '차원 분석',
            'data': dim_result['data'] if dim_result['success'] else [],
            'summary': f"총 {len(dim_result['data'])}개 ALM 차원"
        }

    return report


# ============================================================
# Phase 2: 시나리오 비교 함수
# ============================================================

def compare_scenarios(
    scenario_list: List[int],
    comparison_metrics: Optional[List[str]] = None
) -> Dict[str, Any]:
    """
    여러 시나리오 비교 분석 (Phase 2)

    Args:
        scenario_list: 비교할 시나리오 번호 리스트
        comparison_metrics: 비교할 지표 (None이면 기본 지표)

    Returns:
        {
            'scenarios': List[int],
            'comparison_data': Dict,
            'summary': str
        }
    """
    comparison = {
        'scenarios': scenario_list,
        'comparison_data': {},
        'summary': ''
    }

    for scenario_no in scenario_list:
        query = f"""
        SELECT
            TIME_BAND as 기간대,
            SUM(GAP_PRN_TOTAL) as 원금갭,
            SUM(GAP_INT_TOTAL) as 이자갭,
            SUM(GAP_PRN_TOTAL + GAP_INT_TOTAL) as 총갭
        FROM NFAR_LIQ_GAP_310524
        WHERE SCENARIO_NO = {scenario_no}
        GROUP BY TIME_BAND
        ORDER BY TIME_BAND
        """

        result = execute_sql_query(query)

        if result['success']:
            df = result['dataframe']

            comparison['comparison_data'][f'scenario_{scenario_no}'] = {
                'data': result['data'],
                'total_gap': df['총갭'].sum() if '총갭' in df.columns else 0,
                'max_gap': df['총갭'].max() if '총갭' in df.columns else 0,
                'min_gap': df['총갭'].min() if '총갭' in df.columns else 0,
                'avg_gap': df['총갭'].mean() if '총갭' in df.columns else 0
            }

    summary_lines = []
    summary_lines.append(f"총 {len(scenario_list)}개 시나리오 비교\\n")

    for scenario_no in scenario_list:
        key = f'scenario_{scenario_no}'
        if key in comparison['comparison_data']:
            data = comparison['comparison_data'][key]
            summary_lines.append(
                f"시나리오 {scenario_no}: "
                f"총갭={data['total_gap']:,.0f}, "
                f"평균={data['avg_gap']:,.0f}, "
                f"최대={data['max_gap']:,.0f}, "
                f"최소={data['min_gap']:,.0f}"
            )

    comparison['summary'] = '\\n'.join(summary_lines)
    return comparison


# ============================================================
# Phase 2: 추세 분석 함수
# ============================================================

def analyze_trends(
    metric_type: str,
    currency_or_rate_cd: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
) -> Dict[str, Any]:
    """
    시계열 추세 분석 (Phase 2)

    Args:
        metric_type: 'exchange_rate' 또는 'interest_rate'
        currency_or_rate_cd: 통화 코드 또는 금리 코드
        start_date: 시작 날짜 (YYYY-MM-DD)
        end_date: 종료 날짜 (YYYY-MM-DD)

    Returns:
        {
            'metric_type': str,
            'data_points': List[Dict],
            'statistics': Dict,
            'trend': str
        }
    """
    import numpy as np

    trends = {
        'metric_type': metric_type,
        'data_points': [],
        'statistics': {},
        'trend': ''
    }

    if metric_type == 'exchange_rate':
        query = "SELECT EFFECTIVE_DATE as 일자, EXCH_RATE as 값 FROM NFA_EXCH_RATE_HIST"

        conditions = []
        if currency_or_rate_cd:
            conditions.append(f"UNIT_CURRENCY_CD = '{currency_or_rate_cd}'")
        if start_date:
            conditions.append(f"EFFECTIVE_DATE >= '{start_date}'")
        if end_date:
            conditions.append(f"EFFECTIVE_DATE <= '{end_date}'")

        if conditions:
            query += " WHERE " + " AND ".join(conditions)

        query += " ORDER BY EFFECTIVE_DATE"

    elif metric_type == 'interest_rate':
        query = "SELECT EFFECTIVE_DATE as 일자, INT_RATE as 값 FROM NFA_IRC_RATE_HIST"

        conditions = []
        if currency_or_rate_cd:
            conditions.append(f"INT_RATE_CD = {currency_or_rate_cd}")
        if start_date:
            conditions.append(f"EFFECTIVE_DATE >= '{start_date}'")
        if end_date:
            conditions.append(f"EFFECTIVE_DATE <= '{end_date}'")

        if conditions:
            query += " WHERE " + " AND ".join(conditions)

        query += " ORDER BY EFFECTIVE_DATE"

    else:
        return {
            'error': f"지원하지 않는 metric_type: {metric_type}"
        }

    result = execute_sql_query(query)

    if result['success'] and result['row_count'] > 0:
        df = result['dataframe']
        trends['data_points'] = result['data']

        values = df['값'].values
        trends['statistics'] = {
            'count': len(values),
            'mean': float(np.mean(values)),
            'std': float(np.std(values)),
            'min': float(np.min(values)),
            'max': float(np.max(values)),
            'first_value': float(values[0]),
            'last_value': float(values[-1]),
            'change': float(values[-1] - values[0]),
            'change_pct': float((values[-1] - values[0]) / values[0] * 100) if values[0] != 0 else 0
        }

        if len(values) >= 2:
            x = np.arange(len(values))
            slope = np.polyfit(x, values, 1)[0]

            if slope > 0.01:
                trends['trend'] = '상승 추세'
            elif slope < -0.01:
                trends['trend'] = '하락 추세'
            else:
                trends['trend'] = '안정 추세'

            trends['statistics']['slope'] = float(slope)

    return trends


# ============================================================
# Phase 1,3: 내보내기 함수들
# ============================================================

def export_to_markdown(report_data: Dict[str, Any], output_path: str) -> str:
    """
    리포트를 Markdown 형식으로 내보내기 (Phase 1)
    """
    import os

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)

    lines = []

    # 제목
    lines.append(f"# {report_data['title']}\\n")
    lines.append(f"생성일시: {report_data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')}\\n")

    if report_data['metadata'].get('scenario_no'):
        lines.append(f"시나리오: {report_data['metadata']['scenario_no']}\\n")

    lines.append("\\n---\\n\\n")

    # 각 섹션
    sections = report_data.get('sections', {})

    for section_name, section_data in sections.items():
        lines.append(f"## {section_data['title']}\\n\\n")
        lines.append(f"{section_data.get('summary', '')}\\n\\n")

        # 데이터 테이블
        if 'data' in section_data and section_data['data']:
            data = section_data['data'][:20]
            if data:
                headers = list(data[0].keys())
                lines.append("| " + " | ".join(headers) + " |\\n")
                lines.append("| " + " | ".join(['---'] * len(headers)) + " |\\n")

                for row in data:
                    values = [str(row.get(h, '')) for h in headers]
                    lines.append("| " + " | ".join(values) + " |\\n")

                lines.append("\\n")

        lines.append("\\n")

    with open(output_path, 'w', encoding='utf-8') as f:
        f.writelines(lines)

    return output_path


def export_to_pdf(report_data: Dict[str, Any], output_path: str) -> str:
    """
    리포트를 PDF 형식으로 내보내기 (Phase 3)
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm
    import os

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)

    doc = SimpleDocTemplate(output_path, pagesize=A4)
    story = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30
    )
    heading_style = styles['Heading2']

    story.append(Paragraph(report_data['title'], title_style))
    story.append(Spacer(1, 0.5*cm))

    meta_text = f"생성일시: {report_data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')}"
    if report_data['metadata'].get('scenario_no'):
        meta_text += f" | 시나리오: {report_data['metadata']['scenario_no']}"
    story.append(Paragraph(meta_text, styles['Normal']))
    story.append(Spacer(1, 1*cm))

    sections = report_data.get('sections', {})

    for section_name, section_data in sections.items():
        story.append(Paragraph(section_data['title'], heading_style))
        story.append(Spacer(1, 0.3*cm))

        story.append(Paragraph(section_data.get('summary', ''), styles['Normal']))
        story.append(Spacer(1, 0.5*cm))

        if 'data' in section_data and section_data['data']:
            data = section_data['data'][:10]
            if data:
                headers = list(data[0].keys())
                table_data = [headers]

                for row in data:
                    table_data.append([str(row.get(h, '')) for h in headers])

                t = Table(table_data)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))

                story.append(t)
                story.append(Spacer(1, 0.5*cm))

        story.append(Spacer(1, 0.5*cm))

    doc.build(story)

    return output_path


def export_to_excel(report_data: Dict[str, Any], output_path: str) -> str:
    """
    리포트를 Excel 형식으로 내보내기 (Phase 3)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    import os

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)

    wb = Workbook()

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    summary_ws = wb.create_sheet(title="요약")
    summary_ws.append([report_data['title']])
    summary_ws.append([])
    summary_ws.append(['생성일시', report_data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')])

    if report_data['metadata'].get('scenario_no'):
        summary_ws.append(['시나리오', report_data['metadata']['scenario_no']])

    summary_ws.append([])
    summary_ws.append(['섹션', '요약'])

    title_font = Font(size=16, bold=True)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    summary_ws['A1'].font = title_font

    sections = report_data.get('sections', {})

    for section_name, section_data in sections.items():
        summary_ws.append([section_data['title'], section_data.get('summary', '')])

        ws = wb.create_sheet(title=section_data['title'][:30])

        ws.append([section_data['title']])
        ws.append([section_data.get('summary', '')])
        ws.append([])

        if 'data' in section_data and section_data['data']:
            data = section_data['data']
            if data:
                headers = list(data[0].keys())
                ws.append(headers)

                for col_idx, _ in enumerate(headers, 1):
                    cell = ws.cell(row=ws.max_row, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')

                for row_data in data:
                    ws.append([row_data.get(h, '') for h in headers])

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)

    return output_path


def export_report(
    report_data: Dict[str, Any],
    format: str = 'pdf',
    output_dir: str = './reports'
) -> Dict[str, str]:
    """
    리포트를 지정된 형식으로 내보내기 (Phase 3)
    """
    import os

    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    results = {}

    if format in ['pdf', 'all']:
        pdf_path = os.path.join(output_dir, f'ALM_Report_{timestamp}.pdf')
        export_to_pdf(report_data, pdf_path)
        results['pdf'] = pdf_path

    if format in ['excel', 'all']:
        excel_path = os.path.join(output_dir, f'ALM_Report_{timestamp}.xlsx')
        export_to_excel(report_data, excel_path)
        results['excel'] = excel_path

    if format in ['markdown', 'all']:
        md_path = os.path.join(output_dir, f'ALM_Report_{timestamp}.md')
        export_to_markdown(report_data, md_path)
        results['markdown'] = md_path

    return results

print("✓ Phase 1-3 모든 리포트 함수 정의 완료!")
'''

# Cell 9 마지막에 함수들 추가
if not current_source.endswith('\n'):
    current_source += '\n'

new_source = current_source + all_new_functions

# Cell 9 업데이트
nb['cells'][cell9_idx]['source'] = new_source if isinstance(new_source, list) else [new_source]

# 저장
with open(NOTEBOOK_PATH, 'w') as f:
    json.dump(nb, f, ensure_ascii=False, indent=1)

print("\n" + "="*70)
print("✅ 모든 리포트 함수 추가 완료!")
print("="*70)
print("\n📝 Cell 9에 추가된 함수:")
print("  ✓ generate_comprehensive_report() - Phase 1")
print("  ✓ compare_scenarios() - Phase 2")
print("  ✓ analyze_trends() - Phase 2")
print("  ✓ export_to_markdown() - Phase 1")
print("  ✓ export_to_pdf() - Phase 3")
print("  ✓ export_to_excel() - Phase 3")
print("  ✓ export_report() - Phase 3")
print("\n다음 단계: Cell 12 도구 추가 및 Cell 19 프롬프트 업데이트")
